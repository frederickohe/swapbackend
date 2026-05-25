"""File content extraction utilities for AI training documents.

Supports extracting text from:
- Plain text (.txt)
- PDF files (.pdf)
- Word documents (.docx)
- CSV files (.csv)
"""

import logging
import io
from typing import Optional, Tuple
from fastapi import UploadFile

logger = logging.getLogger(__name__)


class FileContentExtractor:
    """Extract text content from uploaded files."""
    
    # Max content size to extract (100KB) - prevents storing huge files
    MAX_CONTENT_SIZE = 100000  # 100KB
    
    # Supported file types and their MIME types
    SUPPORTED_TYPES = {
        'text/plain': ['.txt'],
        'application/pdf': ['.pdf'],
        'application/vnd.openxmlformats-officedocument.wordprocessingml.document': ['.docx'],
        'text/csv': ['.csv'],
        'application/vnd.ms-excel': ['.xls', '.xlsx'],
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': ['.xlsx'],
    }
    
    @staticmethod
    def extract_content(file: UploadFile, file_content: bytes) -> Optional[str]:
        """
        Extract text content from an uploaded file.
        
        Args:
            file: FastAPI UploadFile object
            file_content: Raw file bytes
            
        Returns:
            Extracted text content or None if extraction failed
        """
        if not file or not file_content:
            return None
        
        content_type = file.content_type or 'application/octet-stream'
        filename = file.filename or ''
        
        try:
            if content_type == 'text/plain' or filename.endswith('.txt'):
                return FileContentExtractor._extract_text(file_content)
            
            elif content_type == 'application/pdf' or filename.endswith('.pdf'):
                return FileContentExtractor._extract_pdf(file_content)
            
            elif content_type in [
                'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                'application/msword'
            ] or filename.endswith(('.docx', '.doc')):
                return FileContentExtractor._extract_docx(file_content)
            
            elif content_type == 'text/csv' or filename.endswith('.csv'):
                return FileContentExtractor._extract_csv(file_content)
            
            elif content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' or filename.endswith('.xlsx'):
                return FileContentExtractor._extract_xlsx(file_content)
            
            else:
                logger.warning(f"Unsupported file type: {content_type} for file {filename}")
                return None
        
        except Exception as e:
            logger.error(f"Error extracting content from {filename}: {str(e)}", exc_info=True)
            return None
    
    @staticmethod
    def _extract_text(file_content: bytes) -> Optional[str]:
        """Extract text from plain text file."""
        try:
            text = file_content.decode('utf-8', errors='replace')
            # Truncate if too large
            if len(text) > FileContentExtractor.MAX_CONTENT_SIZE:
                text = text[:FileContentExtractor.MAX_CONTENT_SIZE] + "\n[Content truncated...]"
            return text.strip() if text.strip() else None
        except Exception as e:
            logger.error(f"Error extracting text: {str(e)}")
            return None
    
    @staticmethod
    def _extract_pdf(file_content: bytes) -> Optional[str]:
        """Extract text from PDF file."""
        try:
            import PyPDF2
            
            pdf_file = io.BytesIO(file_content)
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            
            text_parts = []
            for page_num in range(len(pdf_reader.pages)):
                page = pdf_reader.pages[page_num]
                text_parts.append(page.extract_text())
            
            text = '\n'.join(text_parts)
            
            # Truncate if too large
            if len(text) > FileContentExtractor.MAX_CONTENT_SIZE:
                text = text[:FileContentExtractor.MAX_CONTENT_SIZE] + "\n[Content truncated...]"
            
            return text.strip() if text.strip() else None
            
        except ImportError:
            logger.warning("PyPDF2 not installed, cannot extract PDF content")
            return None
        except Exception as e:
            logger.error(f"Error extracting PDF: {str(e)}")
            return None
    
    @staticmethod
    def _extract_docx(file_content: bytes) -> Optional[str]:
        """Extract text from DOCX file."""
        try:
            from docx import Document
            
            docx_file = io.BytesIO(file_content)
            doc = Document(docx_file)
            
            text_parts = []
            for paragraph in doc.paragraphs:
                if paragraph.text.strip():
                    text_parts.append(paragraph.text)
            
            text = '\n'.join(text_parts)
            
            # Truncate if too large
            if len(text) > FileContentExtractor.MAX_CONTENT_SIZE:
                text = text[:FileContentExtractor.MAX_CONTENT_SIZE] + "\n[Content truncated...]"
            
            return text.strip() if text.strip() else None
            
        except ImportError:
            logger.warning("python-docx not installed, cannot extract DOCX content")
            return None
        except Exception as e:
            logger.error(f"Error extracting DOCX: {str(e)}")
            return None
    
    @staticmethod
    def _extract_csv(file_content: bytes) -> Optional[str]:
        """Extract text from CSV file."""
        try:
            import csv
            
            csv_file = io.StringIO(file_content.decode('utf-8', errors='replace'))
            csv_reader = csv.reader(csv_file)
            
            text_parts = []
            for row in csv_reader:
                text_parts.append(','.join(row))
            
            text = '\n'.join(text_parts)
            
            # Truncate if too large
            if len(text) > FileContentExtractor.MAX_CONTENT_SIZE:
                text = text[:FileContentExtractor.MAX_CONTENT_SIZE] + "\n[Content truncated...]"
            
            return text.strip() if text.strip() else None
            
        except Exception as e:
            logger.error(f"Error extracting CSV: {str(e)}")
            return None
    
    @staticmethod
    def _extract_xlsx(file_content: bytes) -> Optional[str]:
        """Extract text from XLSX file."""
        try:
            import openpyxl
            
            xlsx_file = io.BytesIO(file_content)
            workbook = openpyxl.load_workbook(xlsx_file)
            
            text_parts = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    row_text = '\t'.join(str(cell) if cell is not None else '' for cell in row)
                    if row_text.strip():
                        text_parts.append(row_text)
            
            text = '\n'.join(text_parts)
            
            # Truncate if too large
            if len(text) > FileContentExtractor.MAX_CONTENT_SIZE:
                text = text[:FileContentExtractor.MAX_CONTENT_SIZE] + "\n[Content truncated...]"
            
            return text.strip() if text.strip() else None
            
        except ImportError:
            logger.warning("openpyxl not installed, cannot extract XLSX content")
            return None
        except Exception as e:
            logger.error(f"Error extracting XLSX: {str(e)}")
            return None
